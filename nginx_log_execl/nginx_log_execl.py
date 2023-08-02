import re
from datetime import datetime
import openpyxl
from dateutil.parser import parse
import time
import os
from conf import *


def re_data(data,number): 
    # 日志处理
    log_pattern = r'(.*?) - (-|.*) \[(.*?)\] "(.*?)" (.*?) (.*?) "(.*?|-)" "(.*?)" (.*?)$'
    data_match = re.match(log_pattern, data)
    nginx_log_dict = dict()
    nginx_log_dict["IP"] = data_match.group(1)
    nginx_log_dict["日期"] = data_match.group(3)
    nginx_log_dict["请求方法"] = data_match.group(4)
    nginx_log_dict["请求状态"] = data_match.group(5)
    nginx_log_dict["请求大小"] = data_match.group(6)
    nginx_log_dict["客户端信息"] = data_match.group(8)
    nginx_log_dict["行数"] = number
    # 返回数据字典
    return nginx_log_dict


def time_data(nginx_log_dict):
    # 格式化日期，使其写入到字典内能够更加容易看懂
    # data_match = re_data(data)
    # 定义日期字符串的格式
    date_format = "%d/%b/%Y:%H:%M:%S %z"
    # 将日期字符串解析为datetime对象
    datetime_obj = datetime.strptime(nginx_log_dict['日期'], date_format)
    # 将datetime对象按照指定格式进行格式化
    time_date = datetime_obj.strftime("%Y-%m-%d %H:%M:%S %Z")
    nginx_log_dict['日期'] = time_date
    return nginx_log_dict

def excel_from_name(nginx_log_dict): 
    # 使用数据的日期作为表名
    date_str = nginx_log_dict['日期']
    datetime_obj = parse(date_str)
    # 获取月份和年份，用于创建工作表名称
    month = datetime_obj.strftime("%m")
    year = datetime_obj.strftime("%Y")
    return month, year


def from_header(dict_num,number):
    # 创建execle表格头行数据标题
    print(1)
    workbook_from = openpyxl.load_workbook(data_execl)
    print(2)
    # ws_name = workbook_from.get_sheet_names()
    ws =  workbook_from.create_sheet(str(number))
    print(3)
    ws['A1'] = dict_num[0]
    ws['B1'] = dict_num[1]
    ws['C1'] = dict_num[2]
    ws['D1'] = dict_num[3]
    ws['E1'] = dict_num[4]
    ws['F1'] = dict_num[5]
    ws['G1'] = dict_num[6]
    print(123)
    workbook_from.save(data_execl)



if __name__ == "__main__":
    dict_list = [] # 用于统计字典
    number = 1  # 用于记录行数与列表大小
    workbook = openpyxl.Workbook()
    workbook.save(data_execl)
    with open(file_address,'r') as file:
        # file.seek(log_pointer)
        data = file.readline()
        while data:
            nginx_log_dict = re_data(data,number)
            nginx_log_dict = time_data(nginx_log_dict)
            dict_list.append(nginx_log_dict)
            if len(dict_list) == 10000:
                dict_num = dict_list[0]
                dict_num = list(dict_num.keys())
                # print(type(dict),dict_num)
                from_header(dict_num,number)
                workbook_from = openpyxl.load_workbook(data_execl)
                ws = workbook_from.get_sheet_by_name(str(number))
                for i in dict_list:
                    ws.append(list(i.values()))
                workbook_from.save(data_execl)
                dict_list = []
                time.sleep(10)
            number += 1
            data = file.readline()
            print(number)

            

    

