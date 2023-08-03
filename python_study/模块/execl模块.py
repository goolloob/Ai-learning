# 模块名称openpyxl  下载 pip install openpyxl
import openpyxl # 导入模块

# 用法1
workbook_from = openpyxl.Workbook() # 打开一个execl表格
workbook_from.save('data.xslx') # 保存到这个文件中，这样就创建了一个execl表格文件

# 用法2
workbook_froms = openpyxl.load_workbook('data.xslx') # 打开一个已经存在的execl文件
ws = workbook_froms.create_sheet(str('表名称')) # 创建一张表
ws.append() # 写入数据，可写入列表
ws['A1'] = '' # 单独写入，A1就是第一列的第一行，以此类推
ws = workbook_from.get_sheet_by_name('表名称') # 切换表
workbook_from.get_sheet_names() # 拿到该文件的所有的表名称，是列表形式
ws = workbook_froms.active() # 获取当前活跃的表，也就是当前表格名称
ws.title()# 获取当前表的名称